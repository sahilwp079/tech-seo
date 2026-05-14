"""ReportAgent — generates Excel workbook and AI summary; marks audit done."""

import asyncio
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from agents.base_agent import BaseAgent, AgentResult
from config import settings

C_HEADER = "1E3A5F"
SEV_COL  = {"critical": "E74C3C", "warning": "F39C12", "info": "3498DB"}
PRI_COL  = {"high": "E74C3C", "medium": "F39C12", "low": "27AE60"}


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hfill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)

def _header_row(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = _hfill(C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[1].height = 22

def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


class ReportAgent(BaseAgent):
    name = "ReportAgent"
    dependencies = ["ScoringAgent", "RecommendationAgent"]

    async def run(self) -> AgentResult:
        await self._started()
        try:
            audit   = self.ctx.store.get_audit(self.ctx.audit_id)
            issues  = self.ctx.store.get_issues(self.ctx.audit_id)
            recs    = self.ctx.store.get_recommendations(self.ctx.audit_id)
            wf_log  = self.ctx.store.get_workflow_log(self.ctx.audit_id)
            overall = audit.get("overall_score", -1) if audit else -1
            score_display = overall if overall != -1 else "N/A"

            excel_path = await asyncio.to_thread(
                self._build_excel, issues, recs, wf_log, score_display, self.ctx.base_url
            )

            ai_summary = await self._ai_summary(issues, overall)
            self.ctx.store.update_audit(
                self.ctx.audit_id,
                status="completed",
                completed_at=datetime.utcnow().isoformat(),
                current_agent="",
            )

            await self._emit("audit.completed", {
                "overall_score": overall, "excel_path": excel_path,
                "issues_count": len(issues), "ai_summary": ai_summary,
            })
            await self._completed(score=None, issues_count=0)
            result = AgentResult(
                agent_name=self.name, success=True,
                data={"excel_path": excel_path, "ai_summary": ai_summary},
            )
            return self._validate_contract(result)
        except Exception as exc:
            self.ctx.store.update_audit(self.ctx.audit_id, status="failed", error_message=str(exc))
            await self._failed(str(exc))
            return AgentResult(agent_name=self.name, success=False, error=str(exc))

    def _build_excel(self, issues, recs, wf_log, score_display, base_url: str) -> str:
        wb = Workbook()
        self._sheet_summary(wb.active, score_display, base_url, issues)
        self._sheet_issues(wb.create_sheet("Issues Found"), issues)
        self._sheet_recs(wb.create_sheet("Recommendations"), recs)
        self._sheet_log(wb.create_sheet("Execution Log"), wf_log)

        Path(settings.REPORTS_DIR).mkdir(parents=True, exist_ok=True)
        path = str(Path(settings.REPORTS_DIR) / f"audit_{self.ctx.audit_id}_report.xlsx")
        wb.save(path)
        return path

    def _sheet_summary(self, ws, score, base_url: str, issues: list) -> None:
        ws.title = "Summary"
        ws.merge_cells("A1:D1")
        ws["A1"] = f"SEO Audit — {base_url}   |   Score: {score}/100   |   {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].fill = _hfill(C_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        sev = {"critical": 0, "warning": 0, "info": 0}
        for i in issues:
            sev[i.get("severity", "info")] = sev.get(i.get("severity", "info"), 0) + 1
        for row, (label, val) in enumerate([
            ("Overall Score", f"{score}/100"),
            ("Critical Issues", sev["critical"]),
            ("Warnings", sev["warning"]),
            ("Info", sev["info"]),
            ("Total Issues", len(issues)),
        ], start=3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
        _widths(ws, [30, 20, 30, 20])

    def _sheet_issues(self, ws, issues: list) -> None:
        _header_row(ws, ["Severity", "Agent", "Title", "Page URL", "Section", "Element", "Current Value", "Description", "Fix"])
        _widths(ws, [12, 22, 45, 45, 20, 35, 35, 60, 60])
        ws.freeze_panes = "A2"
        for issue in sorted(issues, key=lambda i: {"critical": 0, "warning": 1, "info": 2}.get(i.get("severity", "info"), 2)):
            ws.append([
                issue.get("severity", ""),
                issue.get("agent_name", ""),
                issue.get("title", ""),
                issue.get("page_url", ""),
                issue.get("section", ""),
                issue.get("element", ""),
                issue.get("current_value", ""),
                issue.get("description", ""),
                issue.get("fix", ""),
            ])
            r = ws.max_row
            ws.cell(r, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(r, 1).fill = _hfill(SEV_COL.get(issue.get("severity", "info"), "888888"))
            for cell in ws[r]:
                cell.border = _border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 45

    def _sheet_recs(self, ws, recs: list) -> None:
        _header_row(ws, ["Priority", "Agent", "Page URL", "Section", "Element", "Action", "Before", "After"])
        _widths(ws, [10, 22, 40, 20, 30, 55, 45, 55])
        ws.freeze_panes = "A2"
        order = {"high": 0, "medium": 1, "low": 2}
        for rec in sorted(recs, key=lambda r: (order.get(r.get("priority", "low"), 2), r.get("page_url", ""))):
            ws.append([
                rec.get("priority", ""),
                rec.get("agent_name", ""),
                rec.get("page_url", ""),
                rec.get("section", ""),
                rec.get("element", ""),
                rec.get("action", ""),
                rec.get("before", ""),
                rec.get("after", ""),
            ])
            r = ws.max_row
            ws.cell(r, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(r, 1).fill = _hfill(PRI_COL.get(rec.get("priority", "low"), "888888"))
            for cell in ws[r]:
                cell.border = _border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 55

    def _sheet_log(self, ws, wf_log: list) -> None:
        _header_row(ws, ["Agent", "Event", "Status", "Score", "Issues", "Timestamp"])
        _widths(ws, [26, 18, 14, 8, 8, 28])
        ws.freeze_panes = "A2"
        for entry in wf_log:
            ws.append([entry.get("agent_name", ""), entry.get("event_type", ""),
                       entry.get("status", ""), entry.get("score", ""),
                       entry.get("issues_count", ""), entry.get("timestamp", "")])
            for cell in ws[ws.max_row]:
                cell.border = _border()
                cell.alignment = Alignment(vertical="center")

    async def _ai_summary(self, issues: list, overall_score: int) -> str:
        from config import settings
        if not settings.GROQ_API_KEY:
            return ""
        critical = [i for i in issues if i.get("severity") == "critical"][:5]
        warnings = [i for i in issues if i.get("severity") == "warning"][:3]
        top = critical + warnings
        lines = "\n".join(f"- [{i['severity'].upper()}] {i['agent_name']}: {i['title']}" for i in top)
        from utils.prompt_manager import prompt_manager
        prompt = prompt_manager.render(
            "report_summary_prompt",
            url=self.ctx.base_url,
            score=overall_score,
            issues_text=lines,
        )
        _usage: dict = {}

        def _call() -> str:
            from groq import Groq
            resp = Groq(api_key=settings.GROQ_API_KEY).chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=200, temperature=0.3,
            )
            if resp.usage:
                _usage["prompt_tokens"]     = resp.usage.prompt_tokens or 0
                _usage["completion_tokens"] = resp.usage.completion_tokens or 0
            return resp.choices[0].message.content or ""

        try:
            from utils.retry import llm_retry
            result = ""
            async for attempt in llm_retry(max_attempts=2):
                with attempt:
                    result = await asyncio.to_thread(_call)
            if _usage:
                from utils.metrics import metrics
                metrics.record_tokens(
                    self.ctx.audit_id,
                    _usage.get("prompt_tokens", 0),
                    _usage.get("completion_tokens", 0),
                )
            return result
        except Exception:
            return ""
